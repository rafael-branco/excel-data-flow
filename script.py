import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import sys


def find_last_non_blank_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0

def get_data_from_specific_rows(filename, sheet_name, row_numbers):
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb[sheet_name]

    rows_data = {}

    for row_num in row_numbers:
        row_data = [cell.value for cell in ws[row_num]]
        rows_data[row_num] = row_data

    return rows_data

def remove_blank_rows_from_bottom(filename, sheet_name):
    
    print("Deleting empty rows from the bottom")
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]

    for row in range(ws.max_row, 0, -1):
        row_data = [ws.cell(row, col).value for col in range(1, 5)]
        if all(cell in (None, "") for cell in row_data):
            ws.delete_rows(row, 1)
    
    wb.save(filename)


def is_row_blank(row):
    for cell in row:
        if cell.value is not None:
            return False
    return True


def remove_blank_rows_from_sheet(filename, sheet_name):

    print("Removing blank rows between data")
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    blank_rows = [row[0].row for row in sheet.iter_rows() if is_row_blank(row)]
    
    blank_rows_sorted = sorted(blank_rows, reverse=True)
    
    for row_num in blank_rows_sorted:
        sheet.delete_rows(row_num)
    
    workbook.save(filename)


def print_progress_bar(iteration, total, prefix='', suffix='', decimals=1, length=50, fill='█', print_end="\r"):
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}')
    sys.stdout.flush()
    if iteration == total: 
        print()


def clear_excel_data(file_name):
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        try:
            worksheet = workbook['ESTOQUE ATUAL']
            if worksheet.max_row > 1:
                worksheet.delete_rows(2, worksheet.max_row - 1)
            workbook.save(file_name)
            print("Data from the 'ESTOQUE ATUAL' sheet has been deleted, leaving only the title.")
        except KeyError:
            print("The 'ESTOQUE ATUAL' sheet was not found in the file.")



def update_stock(file_path_csv, file_path_excel, user_input):

    data_csv = pd.read_csv(file_path_csv, delimiter=';')
    
    workbook = load_workbook(filename=file_path_excel)
    sheet = workbook["ESTOQUE ATUAL"]
    
    next_row = sheet.max_row + 1
    
    print('Copying data from CSV to Excel')
    for row in data_csv.itertuples(index=False):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=next_row, column=col, value=value)
        next_row += 1
    
    print('Formating and updating values')
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column):
        size = row[6].value
        if size in ["P", "M", "G"]:
            row[4].value = "Comum"
        elif size in [46, 48, 50, 52]:
            row[4].value = "Plus"
        row[3].value = user_input
        
        row[7].value = datetime.now().strftime('%d/%m/%Y')    
    workbook.save(filename=file_path_excel)


def delete_rows_in_batches(sheet, rows_to_delete):
    rows_to_delete_sorted = sorted(rows_to_delete, reverse=True)
    
    start_row = None
    end_row = None
    batch_size = 0
    
    for row_num in rows_to_delete_sorted:
        if start_row is None:
            start_row = row_num
            end_row = row_num
            batch_size = 1
        elif row_num == end_row - 1:
            end_row = row_num
            batch_size += 1
        else:
            sheet.delete_rows(end_row, batch_size)
            start_row = row_num
            end_row = row_num
            batch_size = 1
    
    if start_row is not None:
        sheet.delete_rows(end_row, batch_size)


def update_general_stock(input_value, source_file='_ESTOQUE_ATUAL_LOJA.xlsx', target_file='ESTOQUE GERAL.xlsx'):
    print("Loading the target workbook and its sheet")
    target_workbook = load_workbook(target_file)
    target_sheet = target_workbook["ESTOQUE GERAL"]

    print("Collecting rows to delete")
    rows_to_delete = []
    for idx, row in enumerate(target_sheet.iter_rows(min_row=2, values_only=True, max_col=4), start=2):
        if row[3] == int(input_value):
            rows_to_delete.append(idx)

    print("Deleting rows from the bottom")
    delete_rows_in_batches(target_sheet, rows_to_delete)
    
    print("Loading data from source Excel file")
    source_data = pd.read_excel(source_file, sheet_name='ESTOQUE ATUAL', skiprows=1)
    
    print("Appending data to the target sheet")
    for data_row in dataframe_to_rows(source_data, index=False, header=False):
        target_sheet.append(data_row)
    
    print("Saving")
    target_workbook.save(target_file)

    print(f"Updated '{target_file}' successfully with data from '{source_file}'.")


daily_stock_csv = '22.02.2024.csv'
current_stock_excel = '_ESTOQUE_ATUAL_LOJA.xlsx'
current_stock_sheet = 'ESTOQUE ATUAL'
consolidated_stock_excel = 'ESTOQUE GERAL.xlsx'
consolidated_stock_sheet = 'ESTOQUE GERAL'
user_input = int(input("Qual empresa? Digite 1 para LOJA ou 2 para HD: "))

print("\nStarting process...")
clear_excel_data(current_stock_excel)
update_stock(daily_stock_csv, current_stock_excel, user_input)
update_general_stock(user_input, current_stock_excel, consolidated_stock_excel)
remove_blank_rows_from_sheet(consolidated_stock_excel, consolidated_stock_sheet)
remove_blank_rows_from_bottom(consolidated_stock_excel, consolidated_stock_sheet)