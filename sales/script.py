import csv
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re

def process_columns_set_font(file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    # Process columns B, D, and G
    columns_to_process = ['B', 'D', 'G']
    for column in columns_to_process:
        for cell in worksheet[column]:
            # Check if the cell value is a number using a regular expression
            if re.match(r'^-?\d+(?:\.\d+)?$', str(cell.value)):
                # If the cell value is a number, convert it to an integer
                cell.value = int(float(cell.value))
            # Otherwise, leave it as a string (no action needed)

    # Set the font color of the first row to white
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF")

    # Save the workbook
    workbook.save(filename=file_path)


def set_column_h_to_date_format(file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    # Create the named style for dates if it doesn't already exist in the workbook
    date_style_name = "date_style"
    if date_style_name not in workbook.style_names:
        date_style = NamedStyle(name=date_style_name, number_format="DD/MM/YYYY")
        workbook.add_named_style(date_style)
    
    # Apply the style to each non-empty cell in column H
    for cell in worksheet['H']:
        if cell.value:
            cell.style = date_style_name
    
    workbook.save(filename=file_path)




def insert_and_sort_rows(file_path, rows, insert_at_row):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active
    
    if worksheet.tables:
        table = list(worksheet.tables.values())[0]
        _, min_row, _, max_row = range_boundaries(table.ref)
        
        # Insert rows in the middle of the table
        worksheet.insert_rows(insert_at_row, len(rows))
        
        # Insert new data
        for i, row_data in enumerate(rows, start=insert_at_row):
            fields = row_data.split(';')
            for col_num, value in enumerate(fields, start=1):
                worksheet.cell(row=i, column=col_num, value=value)
        
        # Re-reading the table dimensions to include inserted rows
        table_end_row = max_row + len(rows)
        last_col_letter = get_column_letter(worksheet.max_column)
        new_range = f"{table.ref.split(':')[0]}:{last_col_letter}{table_end_row}"
        table.ref = new_range
        
        # Sorting based on the last column
        data = []
        for row in worksheet.iter_rows(min_row=min_row, max_row=table_end_row, values_only=True):
            data.append(row)
        
        last_col_index = worksheet.max_column - 1
        sorted_data = sorted(data[1:], key=lambda x: x[last_col_index])
        
        for i, row_data in enumerate(sorted_data, start=min_row+1):
            for j, value in enumerate(row_data, start=1):
                worksheet.cell(row=i, column=j, value=value)
        
    workbook.save(filename=file_path)



def get_current_date():
    return datetime.now().strftime("%d/%m/%Y")


def read_csv_into_dict(file_path):
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        return {row[reader.fieldnames[0]]: row for row in reader}


def find_rows_only_in_first_csv(file_path1, file_path2):
    csv1_dict = read_csv_into_dict(file_path1)
    csv2_dict = read_csv_into_dict(file_path2)

    only_in_csv1 = {id: csv1_dict[id] for id in csv1_dict if id not in csv2_dict}
    
    return only_in_csv1


previous = "C:\\Users\\User\\Documents\\gitwork\\excel-data-flow\sales\\anterior.csv" #'C:\\Users\\erixy\\OneDrive\\Work\\_Estoques\\anterior.csv'
current =  "C:\\Users\\User\\Documents\\gitwork\\excel-data-flow\sales\\atual.csv"     #'C:\\Users\\erixy\\OneDrive\\Work\\_Estoques\\atual.csv'
sales = "C:\\Users\\User\\Documents\\gitwork\\excel-data-flow\\sales\\VENDAS.xlsx"    #'C:\\Users\\erixy\\OneDrive\\Work\\_Relatórios\\VENDAS.xlsx'

rows_only_in_first_file = find_rows_only_in_first_csv(previous, current)

rows_with_date = []
for id, row in rows_only_in_first_file.items():
    rows_with_date.append(id + str(";") + str(get_current_date()))

insert_and_sort_rows(sales, rows_with_date, 2)
set_column_h_to_date_format(sales)
process_columns_set_font(sales)